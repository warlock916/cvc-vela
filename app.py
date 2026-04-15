from flask import Flask, request, jsonify, send_from_directory, Response, send_file
from flask_cors import CORS
import os, secrets, functools, csv, io, math, hashlib, json, base64
from datetime import date
from werkzeug.utils import secure_filename

app = Flask(__name__, static_folder='static')
CORS(app)

ADMIN_PWD    = os.environ.get('ADMIN_PASSWORD', 'admin123')
DATABASE_URL = os.environ.get('DATABASE_URL', '')
UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

USE_PG = bool(DATABASE_URL)

if USE_PG:
    try:
        import pg8000
        import urllib.parse as _up
        def get_db():
            r = _up.urlparse(DATABASE_URL)
            conn = pg8000.connect(
                host=r.hostname, port=r.port or 5432,
                database=r.path.lstrip('/'),
                user=r.username, password=r.password,
                ssl_context=True
            )
            return conn
        PH = '%s'
    except ImportError:
        USE_PG = False

if not USE_PG:
    import sqlite3
    DB = os.environ.get('DB_PATH', 'cvc.db')
    def get_db():
        conn = sqlite3.connect(DB)
        conn.row_factory = sqlite3.Row
        return conn
    PH = '?'

def rows_to_dicts(rows, cursor=None):
    if not rows: return []
    if USE_PG:
        cols = [desc[0] for desc in cursor.description]
        return [dict(zip(cols, row)) for row in rows]
    return [dict(r) for r in rows]

def row_to_dict(row, cursor=None):
    if row is None: return None
    if USE_PG:
        cols = [desc[0] for desc in cursor.description]
        return dict(zip(cols, row))
    return dict(row)

CRITERI  = ['tec','sen','aff','pro','imp','dis','com']
DAY_KEYS = ['g1','g2','g3','g4','g5','g6','fin']

def hash_pwd(pwd):
    return hashlib.sha256(pwd.encode()).hexdigest()

def init_db():
    voti_cols = [f"{c}_{d} INTEGER" for c in CRITERI for d in DAY_KEYS]
    pts_cols  = [f"pts_{d} INTEGER" for d in DAY_KEYS]
    AI  = "SERIAL" if USE_PG else "INTEGER"
    AIP = "" if USE_PG else "AUTOINCREMENT"
    TS  = "TIMESTAMP DEFAULT NOW()" if USE_PG else "TEXT DEFAULT (datetime('now'))"

    with get_db() as conn:
        cur = conn.cursor()
        cur.execute(f'''CREATE TABLE IF NOT EXISTS turni (
            id         {AI} PRIMARY KEY {AIP},
            numero     INTEGER NOT NULL,
            corso      TEXT NOT NULL,
            pwd_hash   TEXT NOT NULL,
            pwd_plain  TEXT NOT NULL,
            istruttore TEXT NOT NULL,
            created_at {TS},
            UNIQUE(numero, corso)
        )''')
        cur.execute(f'''CREATE TABLE IF NOT EXISTS valutazioni (
            id               {AI} PRIMARY KEY {AIP},
            data             TEXT NOT NULL,
            istruttore       TEXT NOT NULL,
            corso            TEXT NOT NULL,
            turno            INTEGER NOT NULL,
            allievo          TEXT NOT NULL,
            foto_url         TEXT,
            {", ".join(voti_cols)},
            {", ".join(pts_cols)},
            punteggio_finale REAL
        )''')
        cur.execute(f'''CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            tipo  TEXT DEFAULT 'admin',
            turno INTEGER,
            created_at {TS}
        )''')
        conn.commit()

init_db()

PESI = {
    'D1':[.30,.10,.10,.15,.12,.13,.10],'D2':[.30,.10,.10,.15,.12,.13,.10],
    'D3':[.30,.15,.15,.10,.08,.12,.10],'D4':[.30,.15,.20,.05,.08,.12,.10],
    'D5':[.30,.15,.24,.05,.08,.08,.10],'C1':[.35,.15,.10,.10,.10,.10,.10],
    'C2':[.35,.15,.10,.10,.10,.10,.10],'C3':[.30,.20,.15,.05,.10,.10,.10],
    'C4':[.25,.25,.15,.05,.10,.10,.10],'C5':[.25,.25,.15,.05,.10,.10,.10],
}

def calcola_punteggio(corso, voti):
    p = PESI.get(corso)
    if not p or None in voti: return None
    tec,sen,aff,pro,imp,dis,com = voti
    vp = sum(v*w for v,w in zip(voti,p))
    if corso in ('D1','D2','C1','C2'):
        if tec==5 or com==5: return 5
        result = tec+1 if tec+1<vp else vp
    elif corso in ('D3','C3','C4','C5'):
        if tec==5 or sen==5 or com==5: return 5
        lim=min(tec,sen)+1; result=lim if lim<vp else vp
    elif corso=='D5':
        if tec==5 or sen==5 or com==5: return 5
        result=tec+1 if tec+1<vp else vp
    elif corso=='D4':
        if tec==5 or sen==5 or com==5: return 5
        if tec==8 and sen==8: result=max(8,vp)
        else:
            m2=min(tec,sen); result=m2 if min(tec,sen,vp)<m2 else min(tec,sen,vp)
    else: return None
    return int(math.floor(result))

def check_admin(f):
    @functools.wraps(f)
    def wrapper(*args,**kwargs):
        token=request.headers.get('X-Admin-Token','')
        with get_db() as conn:
            cur=conn.cursor()
            cur.execute(f"SELECT token FROM sessions WHERE token={PH} AND tipo='admin'",(token,))
            if not cur.fetchone(): return jsonify({'error':'Non autorizzato'}),401
        return f(*args,**kwargs)
    return wrapper

def check_turno_auth(turno_num, token):
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(f"SELECT token FROM sessions WHERE token={PH} AND (tipo='admin' OR (tipo='turno' AND turno={PH}))",(token,turno_num))
        return cur.fetchone() is not None

# ── Routes base ───────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('static','index.html')

@app.route('/api/login', methods=['POST'])
def login():
    d=request.json or {}
    if d.get('password')!=ADMIN_PWD: return jsonify({'error':'Password errata'}),401
    token=secrets.token_hex(32)
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(f"INSERT INTO sessions(token,tipo) VALUES({PH},{PH})",(token,'admin'))
        conn.commit()
    return jsonify({'token':token,'tipo':'admin'})

@app.route('/api/verify', methods=['GET'])
def verify():
    token=request.headers.get('X-Auth-Token','')
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(f'SELECT tipo,turno FROM sessions WHERE token={PH}',(token,))
        row=cur.fetchone()
    if not row: return jsonify({'valid':False}),401
    tipo=row[0] if USE_PG else row['tipo']
    turno=row[1] if USE_PG else row['turno']
    return jsonify({'valid':True,'tipo':tipo,'turno':turno})

@app.route('/api/turno/<int:numero>/exists', methods=['GET'])
def turno_exists(numero):
    corso=request.args.get('corso','')
    with get_db() as conn:
        cur=conn.cursor()
        if corso: cur.execute(f'SELECT id FROM turni WHERE numero={PH} AND corso={PH}',(numero,corso))
        else: cur.execute(f'SELECT id FROM turni WHERE numero={PH}',(numero,))
        row=cur.fetchone()
    return jsonify({'exists': row is not None})

@app.route('/api/turno/login', methods=['POST'])
def turno_login():
    d=request.json or {}
    numero=d.get('numero'); pwd=d.get('password','').strip()
    istr=d.get('istruttore','').strip(); corso=d.get('corso','').strip()
    if not numero or not pwd: return jsonify({'error':'Turno e password obbligatori'}),400
    try: numero=int(numero)
    except: return jsonify({'error':'Numero turno non valido'}),400
    if not (1<=numero<=60): return jsonify({'error':'Il turno deve essere tra 1 e 60'}),400

    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(f'SELECT * FROM turni WHERE numero={PH} AND corso={PH}',(numero, corso if corso else ''))
        turno_dict=row_to_dict(cur.fetchone(), cur)
        if turno_dict is None:
            if not istr or not corso:
                return jsonify({'error':'Prima apertura: inserisci anche istruttore e corso','primo_accesso':True}),400
            try:
                cur.execute(f'INSERT INTO turni(numero,corso,pwd_hash,pwd_plain,istruttore) VALUES({PH},{PH},{PH},{PH},{PH})',
                           (numero,corso,hash_pwd(pwd),pwd,istr))
                conn.commit()
            except Exception as ex:
                conn.rollback()
                return jsonify({'error':'Errore: '+str(ex)}),500
            cur.execute(f'SELECT * FROM turni WHERE numero={PH} AND corso={PH}',(numero,corso))
            turno_dict=row_to_dict(cur.fetchone(), cur)
        else:
            if hash_pwd(pwd)!=turno_dict['pwd_hash']:
                return jsonify({'error':'Password errata per questo turno'}),401
        token=secrets.token_hex(32)
        cur.execute(f"INSERT INTO sessions(token,tipo,turno) VALUES({PH},{PH},{PH})",(token,'turno',numero))
        conn.commit()
    return jsonify({'token':token,'tipo':'turno','turno':numero,
                    'istruttore':turno_dict['istruttore'],'corso':turno_dict['corso']})

@app.route('/api/turno/<int:numero>', methods=['GET'])
def turno_info(numero):
    token=request.headers.get('X-Auth-Token','')
    if not check_turno_auth(numero,token): return jsonify({'error':'Non autorizzato'}),401
    corso=request.args.get('corso','')
    with get_db() as conn:
        cur=conn.cursor()
        if corso: cur.execute(f'SELECT * FROM turni WHERE numero={PH} AND corso={PH}',(numero,corso))
        else: cur.execute(f'SELECT * FROM turni WHERE numero={PH}',(numero,))
        t=row_to_dict(cur.fetchone(), cur)
        if not t: return jsonify({'error':'Turno non trovato'}),404
        cur.execute(f'SELECT * FROM valutazioni WHERE turno={PH} AND corso={PH} ORDER BY allievo',(numero,t['corso']))
        rows=rows_to_dicts(cur.fetchall(), cur)
    return jsonify({'turno':t,'allievi':rows})

@app.route('/api/scheda', methods=['POST'])
def salva_scheda():
    d=request.json or {}
    records=d.get('records',[]); token=request.headers.get('X-Auth-Token','')
    if not records: return jsonify({'error':'Nessun record'}),400
    oggi=date.today().isoformat(); salvati=0
    with get_db() as conn:
        cur=conn.cursor()
        for rec in records:
            corso=rec.get('corso',''); istr=rec.get('istruttore','').strip()
            allievo=rec.get('allievo','').strip(); turno=rec.get('turno')
            if not all([corso,istr,allievo,turno]): continue
            try: turno=int(turno)
            except: continue
            if not check_turno_auth(turno,token): return jsonify({'error':'Non autorizzato'}),401
            cols,vals=[],[]
            for c in CRITERI:
                for dk in DAY_KEYS:
                    v=rec.get(f'{c}_{dk}')
                    v=int(v) if v is not None and 1<=int(v)<=10 else None
                    cols.append(f'{c}_{dk}'); vals.append(v)
            pts_list=[]
            for dk in DAY_KEYS:
                voti_day=[rec.get(f'{c}_{dk}') for c in CRITERI]
                voti_day=[int(x) if x is not None and 1<=int(x)<=10 else None for x in voti_day]
                pt=calcola_punteggio(corso,voti_day)
                cols.append(f'pts_{dk}'); vals.append(pt); pts_list.append(pt)
            validi=[p for p in pts_list if p is not None]
            pf=int(math.floor(sum(validi)/len(validi))) if validi else None
            cur.execute(f'SELECT id FROM valutazioni WHERE turno={PH} AND allievo={PH} AND corso={PH}',(turno,allievo,corso))
            existing=cur.fetchone()
            if existing:
                eid=existing[0] if USE_PG else existing['id']
                set_clause=','.join(f'{c}={PH}' for c in cols)+f',punteggio_finale={PH}'
                cur.execute(f'UPDATE valutazioni SET {set_clause} WHERE id={PH}',vals+[pf,eid])
            else:
                all_cols=['data','istruttore','corso','turno','allievo']+cols+['punteggio_finale']
                all_vals=[oggi,istr,corso,turno,allievo]+vals+[pf]
                cur.execute(f"INSERT INTO valutazioni ({','.join(all_cols)}) VALUES ({','.join([PH]*len(all_vals))})",all_vals)
            salvati+=1
        conn.commit()
    return jsonify({'ok':True,'salvati':salvati})

# ── Foto allievi ──────────────────────────────────────────────────────────
ALLOWED_EXT = {'png','jpg','jpeg','gif','webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED_EXT

@app.route('/api/foto/<int:turno>/<path:allievo>', methods=['POST'])
def upload_foto(turno, allievo):
    token=request.headers.get('X-Auth-Token','')
    if not check_turno_auth(turno, token): return jsonify({'error':'Non autorizzato'}),401
    if 'foto' not in request.files: return jsonify({'error':'Nessun file'}),400
    file=request.files['foto']
    if not file or not allowed_file(file.filename): return jsonify({'error':'Formato non supportato'}),400

    # Salva con nome sicuro basato su turno+allievo
    ext=file.filename.rsplit('.',1)[1].lower()
    nome_file=secure_filename(f"t{turno}_{allievo}.{ext}")
    path=os.path.join(UPLOAD_FOLDER, nome_file)
    file.save(path)

    # Ridimensiona a 100x100 se PIL disponibile
    try:
        from PIL import Image
        img=Image.open(path)
        img=img.convert('RGB')
        img.thumbnail((100,100), Image.LANCZOS)
        # Crop quadrato centrato
        w,h=img.size
        m=min(w,h)
        img=img.crop(((w-m)//2,(h-m)//2,(w+m)//2,(h+m)//2))
        img=img.resize((100,100), Image.LANCZOS)
        img.save(path, quality=85)
    except: pass

    foto_url=f'/uploads/{nome_file}'
    # Aggiorna DB
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(f'SELECT id FROM valutazioni WHERE turno={PH} AND allievo={PH}',(turno,allievo))
        row=cur.fetchone()
        if row:
            eid=row[0] if USE_PG else row['id']
            cur.execute(f'UPDATE valutazioni SET foto_url={PH} WHERE id={PH}',(foto_url,eid))
            conn.commit()
    return jsonify({'ok':True,'foto_url':foto_url})

@app.route('/api/foto/<int:turno>/<path:allievo>', methods=['DELETE'])
@check_admin
def delete_foto(turno, allievo):
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(f'SELECT foto_url FROM valutazioni WHERE turno={PH} AND allievo={PH}',(turno,allievo))
        row=cur.fetchone()
        if row:
            url=row[0] if USE_PG else row['foto_url']
            if url:
                path=os.path.join(UPLOAD_FOLDER, os.path.basename(url))
                if os.path.exists(path): os.remove(path)
            eid_row=cur.fetchone()
        cur.execute(f'UPDATE valutazioni SET foto_url=NULL WHERE turno={PH} AND allievo={PH}',(turno,allievo))
        conn.commit()
    return jsonify({'ok':True})

@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# ── Riepilogo admin ───────────────────────────────────────────────────────
@app.route('/api/valutazioni/public', methods=['GET'])
@check_admin
def lista_public():
    q=request.args.get('q',''); corso=request.args.get('corso','')
    turno=request.args.get('turno',''); limit=int(request.args.get('limit',500))
    pts_cols=','.join(f'pts_{dk}' for dk in DAY_KEYS)
    where,params=[],[]
    if q:
        like = f'%{q}%'
        if USE_PG: where.append(f'(allievo ILIKE {PH} OR istruttore ILIKE {PH})'); params+=[like,like]
        else: where.append(f'(allievo LIKE {PH} OR istruttore LIKE {PH})'); params+=[like,like]
    if corso: where.append(f'corso={PH}'); params.append(corso)
    if turno: where.append(f'turno={PH}'); params.append(int(turno))
    sql=f'SELECT id,data,istruttore,turno,allievo,corso,foto_url,{pts_cols},punteggio_finale FROM valutazioni'
    csql='SELECT COUNT(*) FROM valutazioni'
    if where:
        w=' WHERE '+' AND '.join(where); sql+=w; csql+=w
    sql+=f' ORDER BY turno,allievo LIMIT {PH}'; params.append(limit)
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(sql,params); rows=rows_to_dicts(cur.fetchall(),cur)
        cur.execute(csql,params[:-1]); total=cur.fetchone()[0]
    return jsonify({'total':total,'rows':rows})

@app.route('/api/valutazioni/<int:vid>', methods=['DELETE'])
@check_admin
def elimina(vid):
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(f'DELETE FROM valutazioni WHERE id={PH}',(vid,)); conn.commit()
    return jsonify({'ok':True})

@app.route('/api/valutazioni/<int:vid>/detail', methods=['GET'])
@check_admin
def detail(vid):
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(f'SELECT * FROM valutazioni WHERE id={PH}',(vid,))
        row=row_to_dict(cur.fetchone(),cur)
    if not row: return jsonify({'error':'Non trovato'}),404
    return jsonify(row)

@app.route('/api/valutazioni/<int:vid>', methods=['PUT'])
@check_admin
def modifica(vid):
    payload=request.json or {}
    sets,vals=[],[]
    for c in CRITERI:
        for dk in DAY_KEYS:
            key=f'{c}_{dk}'
            if key in payload:
                v=payload[key]; v=int(v) if v is not None and 1<=int(v)<=10 else None
                sets.append(f'{key}={PH}'); vals.append(v)
    for dk in DAY_KEYS:
        if f'pts_{dk}' in payload: sets.append(f'pts_{dk}={PH}'); vals.append(payload[f'pts_{dk}'])
    if 'punteggio_finale' in payload: sets.append(f'punteggio_finale={PH}'); vals.append(payload['punteggio_finale'])
    if not sets: return jsonify({'error':'Nessun campo'}),400
    vals.append(vid)
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute(f"UPDATE valutazioni SET {','.join(sets)} WHERE id={PH}",vals); conn.commit()
    return jsonify({'ok':True})

@app.route('/api/stats', methods=['GET'])
@check_admin
def stats():
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute('SELECT COUNT(*) FROM valutazioni'); tot=cur.fetchone()[0]
        cur.execute('SELECT COUNT(DISTINCT istruttore) FROM valutazioni'); istr=cur.fetchone()[0]
        cur.execute('SELECT COUNT(DISTINCT corso) FROM valutazioni'); cors=cur.fetchone()[0]
        cur.execute('SELECT AVG(punteggio_finale) FROM valutazioni WHERE punteggio_finale IS NOT NULL'); med=cur.fetchone()[0]
        cur.execute('SELECT corso,COUNT(*) n,AVG(punteggio_finale) media FROM valutazioni GROUP BY corso ORDER BY corso')
        perc=rows_to_dicts(cur.fetchall(),cur)
        cur.execute('SELECT numero,istruttore,corso,pwd_plain FROM turni ORDER BY numero,corso')
        turni=rows_to_dicts(cur.fetchall(),cur)
    return jsonify({'totale':tot,'istruttori':istr,'corsi_attivi':cors,
                    'media_generale':round(float(med),1) if med else None,
                    'per_corso':perc,'turni':turni})

# ── Export CSV ────────────────────────────────────────────────────────────
@app.route('/api/export/csv', methods=['GET'])
@check_admin
def export_csv():
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute('SELECT * FROM valutazioni ORDER BY turno,allievo')
        rows=cur.fetchall()
        if not rows: return jsonify({'error':'Nessun dato'}),404
        cols=[desc[0] for desc in cur.description]
    out=io.StringIO(); w=csv.writer(out)
    w.writerow(cols)
    for r in rows: w.writerow(list(r))
    return Response('\ufeff'+out.getvalue(),mimetype='text/csv',
                    headers={'Content-Disposition':'attachment; filename=CVC_valutazioni.csv'})

# ── Backup DB (JSON completo) ─────────────────────────────────────────────
@app.route('/api/backup', methods=['GET'])
@check_admin
def backup_db():
    backup={'version':1,'date':date.today().isoformat(),'tables':{}}
    with get_db() as conn:
        cur=conn.cursor()
        for table in ['turni','valutazioni']:
            cur.execute(f'SELECT * FROM {table}')
            rows=cur.fetchall()
            cols=[desc[0] for desc in cur.description]
            backup['tables'][table]=[dict(zip(cols,row)) for row in rows]
    out=json.dumps(backup, ensure_ascii=False, indent=2, default=str)
    return Response(out, mimetype='application/json',
                    headers={'Content-Disposition':f'attachment; filename=CVC_backup_{date.today().isoformat()}.json'})

# ── Restore DB (da JSON) ──────────────────────────────────────────────────
@app.route('/api/restore', methods=['POST'])
@check_admin
def restore_db():
    if 'file' not in request.files: return jsonify({'error':'Nessun file'}),400
    f=request.files['file']
    try: data=json.load(f)
    except: return jsonify({'error':'File JSON non valido'}),400
    if 'tables' not in data: return jsonify({'error':'Formato backup non valido'}),400

    with get_db() as conn:
        cur=conn.cursor()
        # Svuota le tabelle
        cur.execute('DELETE FROM sessions WHERE tipo != %s' if USE_PG else "DELETE FROM sessions WHERE tipo != 'admin'",
                    ('admin',) if USE_PG else ())
        cur.execute('DELETE FROM valutazioni')
        cur.execute('DELETE FROM turni')

        # Reinserisci turni
        for row in data['tables'].get('turni',[]):
            cur.execute(f"INSERT INTO turni(numero,corso,pwd_hash,pwd_plain,istruttore) VALUES({PH},{PH},{PH},{PH},{PH})",
                       (row['numero'],row['corso'],row['pwd_hash'],row['pwd_plain'],row['istruttore']))

        # Reinserisci valutazioni
        voti_cols=[f'{c}_{d}' for c in CRITERI for d in DAY_KEYS]
        pts_cols_=[f'pts_{d}' for d in DAY_KEYS]
        all_cols=['data','istruttore','corso','turno','allievo','foto_url']+voti_cols+pts_cols_+['punteggio_finale']
        for row in data['tables'].get('valutazioni',[]):
            vals=[row.get(c) for c in all_cols]
            cur.execute(f"INSERT INTO valutazioni ({','.join(all_cols)}) VALUES ({','.join([PH]*len(all_cols))})",vals)

        conn.commit()
    return jsonify({'ok':True,'turni':len(data['tables'].get('turni',[])),
                    'valutazioni':len(data['tables'].get('valutazioni',[]))})

# ── Export Excel scheda turno ─────────────────────────────────────────────
@app.route('/api/export/excel/<int:turno>', methods=['GET'])
def export_excel_turno(turno):
    token=request.headers.get('X-Auth-Token','')
    admin_token=request.headers.get('X-Admin-Token','')
    if not check_turno_auth(turno,token) and not admin_token:
        return jsonify({'error':'Non autorizzato'}),401
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error':'openpyxl non installato'}),500

    corso=request.args.get('corso','')
    with get_db() as conn:
        cur=conn.cursor()
        if corso: cur.execute(f'SELECT * FROM turni WHERE numero={PH} AND corso={PH}',(turno,corso))
        else: cur.execute(f'SELECT * FROM turni WHERE numero={PH}',(turno,))
        t=row_to_dict(cur.fetchone(),cur)
        if not t: return jsonify({'error':'Turno non trovato'}),404
        cur.execute(f'SELECT * FROM valutazioni WHERE turno={PH} AND corso={PH} ORDER BY allievo',(turno,t['corso']))
        allievi=rows_to_dicts(cur.fetchall(),cur)

    try:
        wb=Workbook(); ws=wb.active
        ws.title=f'T{turno}_{t["corso"]}'
        BLU=PatternFill('solid',fgColor='1F4E79')
        ORO=PatternFill('solid',fgColor='FFD700')
        DAY_C=['2E75B6','375623','7B5C00','7030A0','C00000','006B6B','8B3A00']
        thin=Side(style='thin',color='CCCCCC')
        BRD=Border(left=thin,right=thin,top=thin,bottom=thin)
        CRITERI_NOMI=['Tecnica','Senso Nautico','Affidabilità','Progressione','Impegno','Disponibilità','Comp. T/I']
        DAY_LABELS=['G1','G2','G3','G4','G5','G6','G7']

        ws.merge_cells('A1:D1')
        ws['A1']='CVC – Blocco Note Volontari'
        ws['A1'].font=Font(bold=True,size=14,color='FFFFFF')
        ws['A1'].fill=BLU
        ws['A1'].alignment=Alignment(horizontal='center')
        ws['E1']=f'Turno: {turno}'; ws['F1']=f'Corso: {t["corso"]}'
        ws['G1']=f'Istruttore: {t["istruttore"]}'; ws['H1']=f'Data: {date.today().isoformat()}'
        ws.row_dimensions[1].height=24

        row=3
        c=ws.cell(row=row,column=1,value='Allievo')
        c.font=Font(bold=True,color='FFFFFF'); c.fill=BLU; c.alignment=Alignment(horizontal='center')
        col=2
        for di,dlbl in enumerate(DAY_LABELS):
            bg=PatternFill('solid',fgColor=DAY_C[di])
            ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+len(CRITERI_NOMI))
            c=ws.cell(row=row,column=col,value=dlbl)
            c.font=Font(bold=True,color='FFFFFF'); c.fill=bg; c.alignment=Alignment(horizontal='center')
            col+=len(CRITERI_NOMI)+1

        row=4
        ws.cell(row=row,column=1,value='').fill=BLU
        col=2
        for di in range(7):
            bg=PatternFill('solid',fgColor=DAY_C[di])
            for crit in CRITERI_NOMI:
                c=ws.cell(row=row,column=col,value=crit[:6])
                c.font=Font(bold=True,color='FFFFFF',size=8); c.fill=bg
                c.alignment=Alignment(horizontal='center',wrap_text=True); c.border=BRD
                col+=1
            c=ws.cell(row=row,column=col,value='Pts')
            c.font=Font(bold=True,color='333333',size=9); c.fill=ORO
            c.alignment=Alignment(horizontal='center'); c.border=BRD
            col+=1

        ws.column_dimensions['A'].width=20
        for i in range(2,col): ws.column_dimensions[get_column_letter(i)].width=7
        ws.row_dimensions[3].height=20; ws.row_dimensions[4].height=30

        for a_row,allievo in enumerate(allievi,start=5):
            ws.row_dimensions[a_row].height=18
            c=ws.cell(row=a_row,column=1,value=allievo['allievo'])
            c.font=Font(bold=True); c.border=BRD
            col=2
            for di,dk in enumerate(['g1','g2','g3','g4','g5','g6','fin']):
                bg=PatternFill('solid',fgColor=DAY_C[di]+'33')
                for key in CRITERI:
                    v=allievo.get(f'{key}_{dk}')
                    c=ws.cell(row=a_row,column=col,value=v)
                    c.fill=bg; c.alignment=Alignment(horizontal='center'); c.border=BRD
                    col+=1
                pts=allievo.get(f'pts_{dk}')
                c=ws.cell(row=a_row,column=col,value=pts)
                c.fill=ORO; c.font=Font(bold=True)
                c.alignment=Alignment(horizontal='center'); c.border=BRD
                col+=1

        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition':f'attachment; filename=CVC_T{turno}_{t["corso"]}.xlsx'}
        )
    except Exception as ex:
        import traceback
        return jsonify({'error':str(ex),'detail':traceback.format_exc()}),500

@app.route('/api/export/pdf/<int:turno>', methods=['GET'])
def export_pdf_turno(turno):
    token=request.headers.get('X-Auth-Token','')
    admin_token=request.headers.get('X-Admin-Token','')
    if not check_turno_auth(turno,token) and not admin_token:
        return jsonify({'error':'Non autorizzato'}),401
    try:
        from reportlab.lib.pagesizes import A4,landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer
        from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
    except ImportError:
        return jsonify({'error':'reportlab non installato'}),500

    corso=request.args.get('corso','')
    with get_db() as conn:
        cur=conn.cursor()
        if corso: cur.execute(f'SELECT * FROM turni WHERE numero={PH} AND corso={PH}',(turno,corso))
        else: cur.execute(f'SELECT * FROM turni WHERE numero={PH}',(turno,))
        t=row_to_dict(cur.fetchone(),cur)
        if not t: return jsonify({'error':'Turno non trovato'}),404
        cur.execute(f'SELECT * FROM valutazioni WHERE turno={PH} AND corso={PH} ORDER BY allievo',(turno,t['corso']))
        allievi=rows_to_dicts(cur.fetchall(),cur)

    try:
        buf=io.BytesIO()
        doc=SimpleDocTemplate(buf,pagesize=landscape(A4),
                              leftMargin=10*mm,rightMargin=10*mm,topMargin=10*mm,bottomMargin=10*mm)
        BLU=colors.HexColor('#1F4E79')
        ORO=colors.HexColor('#FFD700')
        DAY_C=[colors.HexColor(f'#{x}') for x in ['2E75B6','375623','7B5C00','7030A0','C00000','006B6B','8B3A00']]
        CRITERI_NOMI=['Tecnica','Sen.Naut.','Affid.','Progress.','Impegno','Dispon.','Comp.T/I']
        DAY_LABELS=['G1','G2','G3','G4','G5','G6','G7']

        styles=getSampleStyleSheet()
        story=[]
        title_style=ParagraphStyle('t',fontSize=12,textColor=colors.white,backColor=BLU,
                                    spaceAfter=4,alignment=1,fontName='Helvetica-Bold')
        sub_style=ParagraphStyle('s',fontSize=8,spaceAfter=4,fontName='Helvetica')
        story.append(Paragraph('CVC – Blocco Note Volontari',title_style))
        story.append(Paragraph(f'Turno {turno} | Corso {t["corso"]} | Istruttore: {t["istruttore"]} | Data: {date.today().isoformat()}',sub_style))
        story.append(Spacer(1,3*mm))

        header1=['Allievo']
        for lbl in DAY_LABELS:
            header1+=[lbl]+['']*(len(CRITERI_NOMI))
        header2=['']
        for _ in DAY_LABELS:
            header2+=CRITERI_NOMI+['Pts']

        table_data=[header1,header2]
        for allievo in allievi:
            row_data=[allievo['allievo']]
            for dk in ['g1','g2','g3','g4','g5','g6','fin']:
                for key in CRITERI:
                    v=allievo.get(f'{key}_{dk}')
                    row_data.append(str(v) if v is not None else '—')
                pts=allievo.get(f'pts_{dk}')
                row_data.append(str(pts) if pts is not None else '—')
            table_data.append(row_data)

        col_w=[28*mm]+[6*mm]*(7*(len(CRITERI_NOMI)+1))
        tbl=Table(table_data,colWidths=col_w,repeatRows=2)
        style_cmds=[
            ('BACKGROUND',(0,0),(-1,0),BLU),('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,1),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),7),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('ALIGN',(0,2),(0,-1),'LEFT'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),('GRID',(0,0),(-1,-1),0.3,colors.grey),
            ('ROWBACKGROUNDS',(0,2),(-1,-1),[colors.white,colors.HexColor('#F5F5F5')]),
        ]
        col=1
        for di in range(7):
            nc=len(CRITERI_NOMI)+1
            style_cmds+=[
                ('SPAN',(col,0),(col+nc-1,0)),
                ('BACKGROUND',(col,0),(col+nc-1,0),DAY_C[di]),
                ('BACKGROUND',(col,1),(col+nc-2,1),DAY_C[di]),
                ('TEXTCOLOR',(col,1),(col+nc-2,1),colors.white),
                ('BACKGROUND',(col+nc-1,1),(col+nc-1,1),ORO),
                ('TEXTCOLOR',(col+nc-1,1),(col+nc-1,1),colors.black),
            ]
            for r in range(2,len(table_data)):
                style_cmds.append(('BACKGROUND',(col+nc-1,r),(col+nc-1,r),ORO))
            col+=nc
        tbl.setStyle(TableStyle(style_cmds))
        story.append(tbl)
        doc.build(story)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition':f'attachment; filename=CVC_T{turno}_{t["corso"]}.pdf'}
        )
    except Exception as ex:
        import traceback
        return jsonify({'error':str(ex),'detail':traceback.format_exc()}),500

# ── Reset DB ──────────────────────────────────────────────────────────────
@app.route('/api/reset', methods=['POST'])
@check_admin
def reset_db():
    with get_db() as conn:
        cur=conn.cursor()
        cur.execute('DELETE FROM valutazioni')
        cur.execute('DELETE FROM turni')
        cur.execute('DELETE FROM sessions')
        conn.commit()
    return jsonify({'ok':True})

if __name__=='__main__':
    port=int(os.environ.get('PORT',5000))
    app.run(host='0.0.0.0',port=port,debug=False)
